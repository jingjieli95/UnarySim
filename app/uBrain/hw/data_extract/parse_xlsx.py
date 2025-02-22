from openpyxl import load_workbook
import numpy as np
import os.path

class bcolors:
    """
    Reference from: 
    https://svn.blender.org/svnroot/bf-blender/trunk/blender/build_files/scons/tools/bcolors.py
    """
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def get_lists(tab=None, cfg="ubrain"):
    """
    List format = [BUF, RNG, CNT, CMP, PC, REST, TOTAL]
    Units:
        - Area: um^2
        - Power: uW
        - Runtime: ms
        - Energy: nJ
    Returns: 
        area_list, dynamic_list, leakage_list, power_list, runtime, energy_list
    """
    runtime = 14.0/128*10**(3)
    if cfg == 'ubrain':
        area = get_cell_vals(tab, 'G29:G35')
        dynamic = get_cell_vals(tab, 'H29:H35')
        leakage = get_cell_vals(tab, 'I29:I35')
        power = get_cell_vals(tab, 'J29:J35')
        energy = (np.asarray(runtime) * power).tolist()
        return area, dynamic, leakage, power, runtime, energy
    elif cfg == "sc": 
        area = get_cell_vals(tab, 'P29:P35')
        dynamic = get_cell_vals(tab, 'Q29:Q35')
        leakage = get_cell_vals(tab, 'R29:R35')
        power = get_cell_vals(tab, 'S29:S35')
        energy = (np.asarray(runtime) * power).tolist()
        return area, dynamic, leakage, power, runtime, energy
    else:
        print(bcolors.FAIL + f'Unrecognized cfg {cfg}. Options: sc, ubrain' + bcolors.ENDC)
        exit()
    
def get_cell_vals(tab, str_range):
    if ':' in str_range:
        cells = tab[str_range]
        return [x.value for cell in cells for x in cell]
    else:
        cell = tab[str_range]
        return cell.value

def query(str_tab=None, cfg='ubrain'):
    wb = load_workbook(filename="uBrain_resource.xlsx", data_only=True)
    sheets = wb.sheetnames
    ind = sheets.index(str_tab)
    return get_lists(tab=wb[sheets[ind]], cfg=cfg)

def main():
    file_name = "uBrain_resource.xlsx"
    file_exists = os.path.exists(file_name)
    if file_exists == False:
        print(bcolors.FAIL + f'{file_name} does not exist. Did you forget to put it in?' + bcolors.ENDC)
        exit()
    
    # example query
    print(query('conv1-F1', 'ubrain'))
    print(query('conv1-F1', 'sc'))
    print(query('conv1-F2', 'ubrain'))
    print(query('conv1-F2', 'sc'))
    print(query('conv1-F4', 'ubrain'))
    print(query('conv1-F4', 'sc'))
    print(query('conv1-F8', 'ubrain'))
    print(query('conv1-F8', 'sc'))
    print(query('conv1-F16', 'ubrain'))
    print(query('conv1-F16', 'sc'))

    print(query('conv2-F1', 'ubrain'))
    print(query('conv2-F1', 'sc'))
    print(query('conv2-F2', 'ubrain'))
    print(query('conv2-F2', 'sc'))
    print(query('conv2-F4', 'ubrain'))
    print(query('conv2-F4', 'sc'))
    print(query('conv2-F8', 'ubrain'))
    print(query('conv2-F8', 'sc'))
    print(query('conv2-F16', 'ubrain'))
    print(query('conv2-F16', 'sc'))
    print(query('conv2-F32', 'ubrain'))
    print(query('conv2-F32', 'sc'))

    print(query('fc3-F256', 'ubrain'))
    print(query('fc3-F256', 'sc'))

    print(query('rnn4-F1', 'ubrain'))
    print(query('rnn4-F1', 'sc'))

    print(query('fc5-F1', 'ubrain'))
    print(query('fc5-F1', 'sc'))
    
    print(query('fc6-F1', 'ubrain'))
    print(query('fc6-F1', 'sc'))


if __name__ == "__main__":
    main()
